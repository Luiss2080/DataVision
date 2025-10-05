"""
Módulo de exportación a Excel
Funciones para exportar datos, gráficos y análisis a archivos Excel.
"""
import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any, Union
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, ScatterChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.graph_objects as go
import plotly.io as pio
import tempfile
import os
from datetime import datetime


class ExportadorExcel:
    """Clase para exportar análisis completos a Excel con formato profesional."""
    
    def __init__(self, nombre_archivo: str):
        """
        Inicializa el exportador Excel.
        
        Args:
            nombre_archivo: Nombre del archivo Excel a generar
        """
        self.nombre_archivo = nombre_archivo
        self.workbook = Workbook()
        self.hojas = {}
        
        # Eliminar hoja por defecto
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Estilos predefinidos
        self.estilo_encabezado = {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
        }
        
        self.estilo_datos = {
            'font': Font(name='Calibri', size=10),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
        }
        
        self.estilo_numerico = {
            'font': Font(name='Calibri', size=10),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border': Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
        }
    
    def crear_hoja(self, nombre: str, df: Optional[pd.DataFrame] = None) -> Any:
        """
        Crea una nueva hoja de trabajo.
        
        Args:
            nombre: Nombre de la hoja
            df: DataFrame opcional para incluir automáticamente
        
        Returns:
            Objeto worksheet
        """
        # Limpiar nombre de hoja (Excel no permite ciertos caracteres)
        nombre_limpio = self._limpiar_nombre_hoja(nombre)
        
        worksheet = self.workbook.create_sheet(title=nombre_limpio)
        self.hojas[nombre] = worksheet
        
        if df is not None:
            self.agregar_dataframe(worksheet, df, fila_inicio=1, columna_inicio=1)
        
        return worksheet
    
    def _limpiar_nombre_hoja(self, nombre: str) -> str:
        """Limpia el nombre de hoja para que sea válido en Excel."""
        caracteres_invalidos = ['\\', '/', '*', '[', ']', ':', '?']
        nombre_limpio = nombre
        for char in caracteres_invalidos:
            nombre_limpio = nombre_limpio.replace(char, '_')
        return nombre_limpio[:31]  # Excel limita a 31 caracteres
    
    def agregar_dataframe(self, worksheet: Any, df: pd.DataFrame, 
                         fila_inicio: int = 1, columna_inicio: int = 1, 
                         incluir_indice: bool = False) -> Dict[str, int]:
        """
        Agrega un DataFrame a una hoja de trabajo con formato.
        
        Args:
            worksheet: Hoja de trabajo
            df: DataFrame a agregar
            fila_inicio: Fila donde empezar
            columna_inicio: Columna donde empezar
            incluir_indice: Si incluir el índice del DataFrame
        
        Returns:
            Diccionario con posiciones finales
        """
        # Agregar datos
        for r in dataframe_to_rows(df, index=incluir_indice, header=True):
            for c_idx, value in enumerate(r, columna_inicio):
                cell = worksheet.cell(row=fila_inicio, column=c_idx, value=value)
                
                # Aplicar estilos
                if fila_inicio == 1:  # Encabezados
                    cell.font = self.estilo_encabezado['font']
                    cell.fill = self.estilo_encabezado['fill']
                    cell.alignment = self.estilo_encabezado['alignment']
                    cell.border = self.estilo_encabezado['border']
                else:  # Datos
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        cell.font = self.estilo_numerico['font']
                        cell.alignment = self.estilo_numerico['alignment']
                        cell.border = self.estilo_numerico['border']
                        # Formato numérico
                        if isinstance(value, float):
                            cell.number_format = '0.00'
                    else:
                        cell.font = self.estilo_datos['font']
                        cell.alignment = self.estilo_datos['alignment']
                        cell.border = self.estilo_datos['border']
            
            fila_inicio += 1
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas(worksheet, df)
        
        return {
            'fila_final': fila_inicio - 1,
            'columna_final': columna_inicio + len(df.columns) + (1 if incluir_indice else 0) - 1
        }
    
    def _ajustar_ancho_columnas(self, worksheet: Any, df: pd.DataFrame):
        """Ajusta automáticamente el ancho de las columnas."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def agregar_titulo_hoja(self, worksheet: Any, titulo: str, fila: int = 1):
        """Agrega un título destacado a la hoja."""
        worksheet.cell(row=fila, column=1, value=titulo)
        cell = worksheet.cell(row=fila, column=1)
        cell.font = Font(name='Calibri', size=16, bold=True, color='366092')
        cell.alignment = Alignment(horizontal='center')
        
        # Combinar celdas si hay múltiples columnas
        if worksheet.max_column > 1:
            worksheet.merge_cells(
                start_row=fila, start_column=1,
                end_row=fila, end_column=worksheet.max_column
            )
    
    def agregar_metadata(self, worksheet: Any, metadata: Dict[str, str], 
                        fila_inicio: int = 1):
        """Agrega metadatos informativos a la hoja."""
        fila_actual = fila_inicio
        
        for clave, valor in metadata.items():
            worksheet.cell(row=fila_actual, column=1, value=clave)
            worksheet.cell(row=fila_actual, column=2, value=valor)
            
            # Estilo para metadatos
            cell_clave = worksheet.cell(row=fila_actual, column=1)
            cell_clave.font = Font(bold=True)
            
            fila_actual += 1
        
        return fila_actual + 1  # Retornar siguiente fila disponible
    
    def crear_grafico_excel(self, worksheet: Any, tipo_grafico: str, 
                           datos_rango: str, titulo: str = "",
                           posicion: str = "E2", ancho: int = 15, alto: int = 10):
        """
        Crea un gráfico nativo de Excel en la hoja.
        
        Args:
            worksheet: Hoja de trabajo
            tipo_grafico: 'bar', 'line', 'scatter', 'pie'
            datos_rango: Rango de datos (ej: "A1:B10")
            titulo: Título del gráfico
            posicion: Celda donde posicionar el gráfico
            ancho: Ancho del gráfico
            alto: Alto del gráfico
        """
        # Crear referencia a los datos
        data = Reference(worksheet, range_string=datos_rango)
        
        # Seleccionar tipo de gráfico
        if tipo_grafico == 'bar':
            chart = BarChart()
        elif tipo_grafico == 'line':
            chart = LineChart()
        elif tipo_grafico == 'scatter':
            chart = ScatterChart()
        elif tipo_grafico == 'pie':
            chart = PieChart()
        else:
            chart = BarChart()  # Por defecto
        
        # Configurar gráfico
        chart.add_data(data, titles_from_data=True)
        if titulo:
            chart.title = titulo
        
        chart.width = ancho
        chart.height = alto
        
        # Agregar gráfico a la hoja
        worksheet.add_chart(chart, posicion)
    
    def exportar_analisis_completo(self, df: pd.DataFrame, 
                                  analisis_resultados: Dict[str, Any]) -> str:
        """
        Exporta un análisis completo con múltiples hojas.
        
        Args:
            df: DataFrame original
            analisis_resultados: Diccionario con resultados de análisis
        
        Returns:
            Nombre del archivo generado
        """
        # Hoja 1: Datos originales
        hoja_datos = self.crear_hoja("Datos Originales")
        self.agregar_titulo_hoja(hoja_datos, "Dataset Original", 1)
        
        # Metadata
        metadata = {
            "Fecha de análisis": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "Número de filas": f"{len(df):,}",
            "Número de columnas": str(len(df.columns)),
            "Memoria utilizada (MB)": f"{df.memory_usage(deep=True).sum() / 1024**2:.2f}"
        }
        fila_siguiente = self.agregar_metadata(hoja_datos, metadata, 3)
        
        # Agregar datos (muestra limitada para Excel)
        df_muestra = df.head(1000)  # Limitar a 1000 filas para rendimiento
        self.agregar_dataframe(hoja_datos, df_muestra, fila_siguiente + 1)
        
        # Hoja 2: Estadísticas descriptivas
        if 'estadisticas' in analisis_resultados:
            hoja_stats = self.crear_hoja("Estadísticas")
            self.agregar_titulo_hoja(hoja_stats, "Estadísticas Descriptivas", 1)
            self.agregar_dataframe(hoja_stats, analisis_resultados['estadisticas'], 3)
        
        # Hoja 3: Análisis de calidad
        if 'calidad_datos' in analisis_resultados:
            hoja_calidad = self.crear_hoja("Calidad de Datos")
            self.agregar_titulo_hoja(hoja_calidad, "Análisis de Calidad", 1)
            self.agregar_dataframe(hoja_calidad, analisis_resultados['calidad_datos'], 3)
        
        # Hoja 4: Correlaciones
        if 'correlaciones' in analisis_resultados:
            hoja_corr = self.crear_hoja("Correlaciones")
            self.agregar_titulo_hoja(hoja_corr, "Análisis de Correlaciones", 1)
            
            # Matriz de correlación
            matriz_corr = analisis_resultados['correlaciones']['matriz_correlacion']
            if matriz_corr is not None:
                self.agregar_dataframe(hoja_corr, matriz_corr, 3)
            
            # Correlaciones significativas
            if 'correlaciones_significativas' in analisis_resultados['correlaciones']:
                corr_sig = pd.DataFrame(analisis_resultados['correlaciones']['correlaciones_significativas'])
                if not corr_sig.empty:
                    posiciones = self.agregar_dataframe(hoja_corr, matriz_corr, 3)
                    fila_sig = posiciones['fila_final'] + 3
                    
                    hoja_corr.cell(row=fila_sig, column=1, value="Correlaciones Significativas")
                    cell = hoja_corr.cell(row=fila_sig, column=1)
                    cell.font = Font(bold=True, size=12)
                    
                    self.agregar_dataframe(hoja_corr, corr_sig, fila_sig + 1)
        
        # Hoja 5: Resumen ejecutivo
        hoja_resumen = self.crear_hoja("Resumen Ejecutivo")
        self.agregar_titulo_hoja(hoja_resumen, "Resumen del Análisis", 1)
        
        # Crear resumen textual
        resumen_datos = []
        
        # Información básica
        resumen_datos.append(["Métrica", "Valor", "Descripción"])
        resumen_datos.append(["Filas totales", f"{len(df):,}", "Número de observaciones"])
        resumen_datos.append(["Columnas totales", len(df.columns), "Número de variables"])
        resumen_datos.append(["Valores nulos", f"{df.isnull().sum().sum():,}", "Total de valores faltantes"])
        resumen_datos.append(["% Completitud", f"{((df.count().sum() / (len(df) * len(df.columns))) * 100):.1f}%", "Porcentaje de datos completos"])
        
        # Tipos de datos
        tipos_count = df.dtypes.value_counts()
        for dtype, count in tipos_count.items():
            resumen_datos.append([f"Columnas {dtype}", count, f"Variables de tipo {dtype}"])
        
        # Crear DataFrame del resumen
        df_resumen = pd.DataFrame(resumen_datos[1:], columns=resumen_datos[0])
        self.agregar_dataframe(hoja_resumen, df_resumen, 3)
        
        # Guardar archivo
        self.workbook.save(self.nombre_archivo)
        return self.nombre_archivo
    
    def cerrar(self):
        """Cierra el libro de trabajo y guarda cambios."""
        if hasattr(self, 'workbook'):
            self.workbook.save(self.nombre_archivo)


def exportar_dataframe_simple(df: pd.DataFrame, nombre_archivo: str, 
                             nombre_hoja: str = "Datos") -> str:
    """
    Exporta un DataFrame simple a Excel con formato básico.
    
    Args:
        df: DataFrame a exportar
        nombre_archivo: Nombre del archivo Excel
        nombre_hoja: Nombre de la hoja
    
    Returns:
        Nombre del archivo generado
    """
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=nombre_hoja, index=False)
        
        # Obtener workbook y worksheet para formato
        workbook = writer.book
        worksheet = writer.sheets[nombre_hoja]
        
        # Formatear encabezados
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return nombre_archivo


def exportar_multiples_dataframes(dataframes_dict: Dict[str, pd.DataFrame], 
                                 nombre_archivo: str) -> str:
    """
    Exporta múltiples DataFrames a diferentes hojas de un archivo Excel.
    
    Args:
        dataframes_dict: Diccionario con {nombre_hoja: dataframe}
        nombre_archivo: Nombre del archivo Excel
    
    Returns:
        Nombre del archivo generado
    """
    exportador = ExportadorExcel(nombre_archivo)
    
    for nombre_hoja, df in dataframes_dict.items():
        hoja = exportador.crear_hoja(nombre_hoja)
        exportador.agregar_titulo_hoja(hoja, nombre_hoja.replace('_', ' ').title(), 1)
        exportador.agregar_dataframe(hoja, df, 3)
    
    exportador.cerrar()
    return nombre_archivo


def exportar_con_graficos_plotly(df: pd.DataFrame, graficos: Dict[str, go.Figure], 
                                nombre_archivo: str) -> str:
    """
    Exporta datos y gráficos de Plotly a Excel (gráficos como imágenes).
    
    Args:
        df: DataFrame con los datos
        graficos: Diccionario con gráficos de Plotly
        nombre_archivo: Nombre del archivo Excel
    
    Returns:
        Nombre del archivo generado
    """
    # Crear archivo base con datos
    exportador = ExportadorExcel(nombre_archivo)
    
    # Hoja de datos
    hoja_datos = exportador.crear_hoja("Datos", df)
    
    # Hoja de gráficos (como imágenes)
    if graficos:
        hoja_graficos = exportador.crear_hoja("Gráficos")
        exportador.agregar_titulo_hoja(hoja_graficos, "Visualizaciones", 1)
        
        fila_actual = 3
        for nombre, fig in graficos.items():
            try:
                # Exportar gráfico como imagen temporal
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
                    pio.write_image(fig, tmp_file.name, width=800, height=600)
                    
                    # Agregar título del gráfico
                    hoja_graficos.cell(row=fila_actual, column=1, value=f"Gráfico: {nombre}")
                    cell = hoja_graficos.cell(row=fila_actual, column=1)
                    cell.font = Font(bold=True, size=12)
                    
                    # Nota: En esta versión simplificada, no insertamos la imagen
                    # Se puede implementar usando openpyxl.drawing.image.Image
                    fila_actual += 2
                    
                    os.unlink(tmp_file.name)
            except Exception as e:
                print(f"Error al procesar gráfico {nombre}: {str(e)}")
    
    exportador.cerrar()
    return nombre_archivo